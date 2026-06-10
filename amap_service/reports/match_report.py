"""线路「原始轨迹长度」对比「匹配成路段总长度」的报表。

- 原始轨迹长度：transit_line_raw 里每条线最新一条响应,经 parse_line_tracks 取各方向
  LineLonLat,算折线总长(米)。
- 匹配后长度：transit_segment 各方向所有 seg 的 line_track(按行进方向裁剪后的路段)折线长之和(米)。
- 差异百分比：(匹配后 - 原始) / 原始 * 100。
"""
import csv
import json
import logging
import os

from sqlalchemy import Engine, func, select

from amap_service.db.schema import transit_line_raw, transit_segment
from amap_service.parsing.transit import parse_line_tracks
from amap_service.sdk import geometry

logger = logging.getLogger(__name__)

_DIR_LABEL = {0: "上行", 1: "下行"}  # schema: 0=上行/单环, 1=下行
CSV_HEADERS = ["线路名称", "上下行", "原始轨迹长度", "匹配后长度", "差异百分比"]


def _polyline_length_m(track: str) -> float:
    pts = geometry.parse_track(track or "")
    return sum(geometry.haversine(pts[i], pts[i + 1]) for i in range(len(pts) - 1))


def _matched_lengths(engine: Engine) -> dict:
    """(line_name, direction) -> 匹配后折线总长(米)，按 seq 拼接各 seg。"""
    with engine.connect() as conn:
        rows = conn.execute(
            select(transit_segment.c.line_name, transit_segment.c.direction,
                   transit_segment.c.line_track)
            .order_by(transit_segment.c.line_name, transit_segment.c.direction,
                      transit_segment.c.seq)
        ).all()
    out: dict = {}
    for r in rows:
        key = (r.line_name, r.direction)
        out[key] = out.get(key, 0.0) + _polyline_length_m(r.line_track)
    return out


def _original_lengths(engine: Engine) -> dict:
    """(line_name, direction) -> 原始 LineLonLat 折线长(米)，每条线取最新一条 raw(max id)。"""
    with engine.connect() as conn:
        latest = (
            select(func.max(transit_line_raw.c.id).label("mid"))
            .group_by(transit_line_raw.c.line_name)
            .subquery()
        )
        rows = conn.execute(
            select(transit_line_raw.c.raw_response)
            .join(latest, transit_line_raw.c.id == latest.c.mid)
        ).all()
    out: dict = {}
    for r in rows:
        try:
            parsed = json.loads(r.raw_response)
        except (TypeError, ValueError):
            continue
        for t in parse_line_tracks(parsed):
            out[(str(t["line_name"]), t["direction"])] = _polyline_length_m(t["track"])
    return out


def build_match_report(engine: Engine) -> list[dict]:
    """每个有匹配路段的 (线路, 方向) 一行。原始轨迹缺失时 original_len_m/diff_pct 为 None。"""
    matched = _matched_lengths(engine)
    original = _original_lengths(engine)
    report = []
    for line_name, direction in sorted(matched, key=lambda k: (str(k[0]), k[1])):
        m = matched[(line_name, direction)]
        o = original.get((line_name, direction))
        report.append({
            "line_name": line_name,
            "direction": direction,
            "direction_label": _DIR_LABEL.get(direction, str(direction)),
            "original_len_m": round(o, 1) if o is not None else None,
            "matched_len_m": round(m, 1),
            "diff_pct": round((m - o) / o * 100, 2) if o else None,
        })
    return report


def _row_cells(r: dict) -> list:
    return [
        r["line_name"], r["direction_label"],
        "" if r["original_len_m"] is None else r["original_len_m"],
        r["matched_len_m"],
        "" if r["diff_pct"] is None else r["diff_pct"],
    ]


def write_match_report_csv(rows: list[dict], path: str) -> None:
    """写 CSV(utf-8-sig,Excel 直接识别中文)。原始/差异缺失留空。"""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADERS)
        for r in rows:
            w.writerow(_row_cells(r))


def _diff_fill(diff_pct):
    """按差异绝对值给整行底色：>10% 红，5%~10%(含10) 黄，其余/缺失 无。"""
    from openpyxl.styles import PatternFill
    if diff_pct is None:
        return None
    a = abs(diff_pct)
    if a > 10:
        return PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    if a >= 5:
        return PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    return None


def write_match_report_xlsx(rows: list[dict], path: str) -> None:
    """写 XLSX：差异绝对值 >10% 整行标红、5%~10% 标黄；原始/差异缺失留空、不标色。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "匹配长度对比"
    ws.append(CSV_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append(_row_cells(r))
        fill = _diff_fill(r["diff_pct"])
        if fill is not None:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    for i, width in enumerate((12, 8, 14, 14, 12), start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"  # 冻结表头
    wb.save(path)
