import json

from sqlalchemy import insert

from amap_service.config.schema import DatabaseConfig, SqliteConfig
from amap_service.db.engine import make_engine
from amap_service.db.migrate import init_db
from amap_service.db.schema import transit_segment, transit_line_raw
from amap_service.sdk import geometry
from amap_service.reports.match_report import (
    build_match_report, write_match_report_csv, write_match_report_xlsx,
)


def _eng(tmp_path):
    eng = make_engine(DatabaseConfig(type="sqlite", sqlite=SqliteConfig(path=str(tmp_path / "t.db"))))
    init_db(eng)
    return eng


def _plen(s):
    pts = geometry.parse_track(s)
    return sum(geometry.haversine(pts[i], pts[i + 1]) for i in range(len(pts) - 1))


def test_build_match_report_computes_lengths_and_diff(tmp_path):
    eng = _eng(tmp_path)
    seg0 = "121.0,31.0;121.001,31.0"
    seg1 = "121.001,31.0;121.002,31.0"
    orig = "121.0,31.0;121.0005,31.0005;121.001,31.0;121.002,31.0"  # 带绕行,比匹配后长
    with eng.begin() as c:
        c.execute(insert(transit_segment), [
            {"line_name": "47", "direction": 0, "seq": 0, "link_id": 1,
             "reverse_coords": 0, "line_track": seg0},
            {"line_name": "47", "direction": 0, "seq": 1, "link_id": 2,
             "reverse_coords": 0, "line_track": seg1}])
        c.execute(insert(transit_line_raw), [
            {"line_name": "47", "raw_response": json.dumps(
                {"Data": {"LineName": "47", "UpObject": {"UpDown": 0, "LineLonLat": orig}}})}])
    rows = build_match_report(eng)
    assert len(rows) == 1
    r = rows[0]
    assert r["line_name"] == "47"
    assert r["direction"] == 0 and r["direction_label"] == "上行"
    matched, original = _plen(seg0) + _plen(seg1), _plen(orig)
    assert abs(r["matched_len_m"] - round(matched, 1)) < 0.05
    assert abs(r["original_len_m"] - round(original, 1)) < 0.05
    assert r["diff_pct"] == round((matched - original) / original * 100, 2)


def test_missing_original_yields_none(tmp_path):
    eng = _eng(tmp_path)
    with eng.begin() as c:
        c.execute(insert(transit_segment), [
            {"line_name": "99", "direction": 1, "seq": 0, "link_id": 1,
             "reverse_coords": 0, "line_track": "121.0,31.0;121.001,31.0"}])
    rows = build_match_report(eng)
    assert rows[0]["direction_label"] == "下行"
    assert rows[0]["original_len_m"] is None
    assert rows[0]["diff_pct"] is None


def test_latest_raw_per_line_used(tmp_path):
    eng = _eng(tmp_path)
    with eng.begin() as c:
        c.execute(insert(transit_segment), [
            {"line_name": "47", "direction": 0, "seq": 0, "link_id": 1,
             "reverse_coords": 0, "line_track": "121.0,31.0;121.001,31.0"}])
        # 旧响应(短) + 新响应(长) → 取最新(max id)
        c.execute(insert(transit_line_raw), [
            {"line_name": "47", "raw_response": json.dumps(
                {"Data": {"LineName": "47", "UpObject": {"UpDown": 0, "LineLonLat": "121.0,31.0;121.0005,31.0"}}})},
            {"line_name": "47", "raw_response": json.dumps(
                {"Data": {"LineName": "47", "UpObject": {"UpDown": 0, "LineLonLat": "121.0,31.0;121.002,31.0"}}})}])
    rows = build_match_report(eng)
    assert abs(rows[0]["original_len_m"] - round(_plen("121.0,31.0;121.002,31.0"), 1)) < 0.05


def test_write_csv_headers_and_values(tmp_path):
    rows = [{"line_name": "47", "direction": 0, "direction_label": "上行",
             "original_len_m": 100.0, "matched_len_m": 110.0, "diff_pct": 10.0},
            {"line_name": "99", "direction": 1, "direction_label": "下行",
             "original_len_m": None, "matched_len_m": 50.0, "diff_pct": None}]
    out = tmp_path / "r.csv"
    write_match_report_csv(rows, str(out))
    text = out.read_text(encoding="utf-8-sig")
    lines = text.strip().splitlines()
    assert lines[0] == "线路名称,上下行,原始轨迹长度,匹配后长度,差异百分比"
    assert lines[1] == "47,上行,100.0,110.0,10.0"
    assert lines[2] == "99,下行,,50.0,"   # 原始/差异缺失留空


def test_write_xlsx_conditional_fills(tmp_path):
    from openpyxl import load_workbook
    rows = [
        {"line_name": "A", "direction": 0, "direction_label": "上行",
         "original_len_m": 100.0, "matched_len_m": 112.0, "diff_pct": 12.0},   # >10 红
        {"line_name": "B", "direction": 0, "direction_label": "上行",
         "original_len_m": 100.0, "matched_len_m": 93.0, "diff_pct": -7.0},     # |7| 黄
        {"line_name": "C", "direction": 0, "direction_label": "上行",
         "original_len_m": 100.0, "matched_len_m": 102.0, "diff_pct": 2.0},     # <5 无色
        {"line_name": "D", "direction": 1, "direction_label": "下行",
         "original_len_m": None, "matched_len_m": 50.0, "diff_pct": None},      # 缺失 无色
        {"line_name": "E", "direction": 0, "direction_label": "上行",
         "original_len_m": 100.0, "matched_len_m": 90.0, "diff_pct": -10.0},    # 边界10 黄
    ]
    out = tmp_path / "r.xlsx"
    write_match_report_xlsx(rows, str(out))
    ws = load_workbook(str(out)).active
    assert [c.value for c in ws[1]] == ["线路名称", "上下行", "原始轨迹长度", "匹配后长度", "差异百分比"]

    def fill_rgb(row):
        return ws.cell(row=row, column=5).fill.start_color.rgb

    assert fill_rgb(2) == "FFFF0000"   # A 红(>10)
    assert fill_rgb(3) == "FFFFFF00"   # B 黄(|7|)
    assert ws.cell(row=4, column=5).fill.fill_type in (None, "none")  # C 无色
    assert ws.cell(row=5, column=5).fill.fill_type in (None, "none")  # D 缺失无色
    assert fill_rgb(6) == "FFFFFF00"   # E 边界10 → 黄
    # 值与留空
    assert ws.cell(row=2, column=1).value == "A"
    assert ws.cell(row=5, column=3).value in ("", None)  # D 原始留空


def test_write_match_report_db_replaces(tmp_path):
    from sqlalchemy import func, select
    from amap_service.db.schema import transit_match_report
    from amap_service.reports.match_report import write_match_report_db
    eng = _eng(tmp_path)  # init_db 已建 transit_match_report 表
    write_match_report_db(eng, [
        {"line_name": "47", "direction": 0, "direction_label": "上行",
         "original_len_m": 100.0, "matched_len_m": 110.0, "diff_pct": 10.0},
        {"line_name": "99", "direction": 1, "direction_label": "下行",
         "original_len_m": None, "matched_len_m": 50.0, "diff_pct": None}])
    with eng.connect() as c:
        got = c.execute(
            select(transit_match_report.c.line_name, transit_match_report.c.direction,
                   transit_match_report.c.original_length_m, transit_match_report.c.matched_length_m,
                   transit_match_report.c.diff_pct).order_by(transit_match_report.c.line_name)
        ).all()
    assert [tuple(r) for r in got] == [("47", 0, 100.0, 110.0, 10.0), ("99", 1, None, 50.0, None)]

    # 整体替换：第二次跑只剩新数据
    write_match_report_db(eng, [
        {"line_name": "47", "direction": 0, "direction_label": "上行",
         "original_len_m": 1.0, "matched_len_m": 2.0, "diff_pct": 100.0}])
    with eng.connect() as c:
        cnt = c.execute(select(func.count()).select_from(transit_match_report)).scalar()
    assert cnt == 1
